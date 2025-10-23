import os
from flask import Blueprint, render_template, request, redirect, url_for, flash, current_app, send_file
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
from datetime import datetime
from werkzeug.utils import secure_filename

from . import db
from .models import Upload, Criteria, Subcriteria
from .forms import UploadForm, CriteriaForm, SubcriteriaForm
from .utils import preprocess_dataframe, roc_weights, aras_compute, aras_score_and_rank

main_bp = Blueprint("main", __name__)

ALLOWED_EXT = {".xlsx", ".xls"}

def _latest_upload():
    return Upload.query.order_by(Upload.created_at.desc()).first()

@main_bp.route("/")
def index():
    return redirect(url_for("main.dashboard"))

@main_bp.route("/dashboard")
@login_required
def dashboard():
    up = _latest_upload()
    return render_template("dashboard.html", upload=up)

# ---------------- Upload data ----------------
@main_bp.route("/upload", methods=["GET","POST"])
@login_required
def upload():
    form = UploadForm()
    if form.validate_on_submit():
        f = form.file.data
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_EXT:
            flash("Format harus Excel (.xlsx / .xls).", "warning")
            return redirect(url_for("main.upload"))
        safe = secure_filename(f.filename)
        save_path = os.path.join(current_app.config["UPLOAD_FOLDER"], safe)
        f.save(save_path)

        # Process
        try:
            df = pd.read_excel(save_path)
        except Exception as e:
            flash(f"Gagal membaca Excel: {e}", "danger")
            return redirect(url_for("main.upload"))

        df = preprocess_dataframe(df)
        processed_path = save_path + ".processed.csv"
        df.to_csv(processed_path, index=False)

        up = Upload(
            user_id=current_user.id,
            original_filename=safe,
            saved_path=save_path,
            processed_path=processed_path,
            n_rows=len(df),
        )
        db.session.add(up)
        db.session.commit()
        flash("Upload & proses berhasil.", "success")
        return redirect(url_for("main.upload"))
    uploads = Upload.query.order_by(Upload.created_at.desc()).all()
    return render_template("upload.html", form=form, uploads=uploads)

@main_bp.route("/upload/<int:uid>/view")
@login_required
def upload_view(uid):
    """
    Read detail: tampilkan pratinjau data mentah & data olah.
    """
    up = Upload.query.get_or_404(uid)

    raw_html = "<em>File tidak ditemukan.</em>"
    proc_html = "<em>Belum ada data olah.</em>"

    # Raw (Excel)
    if up.saved_path and os.path.exists(up.saved_path):
        try:
            # batasi preview 100 baris
            dfr = pd.read_excel(up.saved_path).head(100)
            raw_html = dfr.to_html(classes="table table-sm table-striped", index=False)
        except Exception as e:
            raw_html = f"<em>Gagal baca raw: {e}</em>"

    # Processed (CSV)
    if up.processed_path and os.path.exists(up.processed_path):
        try:
            dfp = pd.read_csv(up.processed_path).head(100)
            proc_html = dfp.to_html(classes="table table-sm table-striped", index=False)
        except Exception as e:
            proc_html = f"<em>Gagal baca processed: {e}</em>"

    return render_template(
        "upload_view.html",
        up=up,
        raw_html=raw_html,
        proc_html=proc_html,
    )

@main_bp.route("/upload/<int:uid>/reprocess", methods=["POST"])
@login_required
def upload_reprocess(uid):
    """
    Update: reprocess raw -> processed.csv ulang.
    """
    up = Upload.query.get_or_404(uid)
    if not up.saved_path or not os.path.exists(up.saved_path):
        flash("File mentah tidak ada. Silakan upload ulang.", "warning")
        return redirect(url_for("main.upload"))

    try:
        df = pd.read_excel(up.saved_path)
        df = preprocess_dataframe(df)
        processed_path = up.saved_path + ".processed.csv"
        df.to_csv(processed_path, index=False)
        up.processed_path = processed_path
        up.n_rows = len(df)
        db.session.commit()
        flash("Reproses berhasil.", "success")
    except Exception as e:
        flash(f"Gagal memproses ulang: {e}", "danger")
    return redirect(url_for("main.upload_view", uid=uid))

@main_bp.route("/upload/<int:uid>/delete", methods=["POST"])
@login_required
def upload_delete(uid):
    """
    Delete: hapus record + file fisik (raw & processed).
    """
    up = Upload.query.get_or_404(uid)
    # hapus file fisik
    for p in [up.saved_path, up.processed_path]:
        if p and os.path.exists(p):
            try:
                os.remove(p)
            except Exception:
                pass
    db.session.delete(up)
    db.session.commit()
    flash("Upload dihapus.", "info")
    return redirect(url_for("main.upload"))

# Download RAW dari folder uploads (tambahan route)
@main_bp.route("/uploads/<path:filename>")
@login_required
def download_raw(filename):
    path = os.path.join(current_app.config["UPLOAD_FOLDER"], filename)
    if not os.path.exists(path):
        flash("File tidak ditemukan.", "warning")
        return redirect(url_for("main.upload"))
    return send_file(path, as_attachment=True)

# ----------------- Criteria CRUD + ROC -----------------
@main_bp.route("/criteria", methods=["GET","POST"])
@login_required
def criteria_list():
    form = CriteriaForm()
    if form.validate_on_submit():
        c = Criteria(name=form.name.data.strip(), ctype=form.ctype.data, display_order=form.display_order.data)
        db.session.add(c)
        db.session.commit()
        flash("Kriteria ditambahkan.", "success")
        return redirect(url_for("main.criteria_list"))
    items = Criteria.query.order_by(Criteria.display_order.asc()).all()
    # ROC weights
    n = len(items)
    weights = {}
    if n > 0:
        roc = roc_weights(n)
        for i, c in enumerate(items):
            weights[c.name] = roc[i]
    return render_template("criteria.html", form=form, items=items, weights=weights)

@main_bp.route("/criteria/<int:cid>/edit", methods=["GET","POST"])
@login_required
def criteria_edit(cid):
    c = Criteria.query.get_or_404(cid)
    form = CriteriaForm(obj=c)
    if form.validate_on_submit():
        c.name = form.name.data.strip()
        c.ctype = form.ctype.data
        c.display_order = form.display_order.data
        db.session.commit()
        flash("Kriteria diupdate.", "success")
        return redirect(url_for("main.criteria_list"))
    return render_template("criteria_edit.html", form=form, c=c)

@main_bp.route("/criteria/<int:cid>/delete", methods=["POST"])
@login_required
def criteria_delete(cid):
    c = Criteria.query.get_or_404(cid)
    db.session.delete(c)
    db.session.commit()
    flash("Kriteria dihapus.", "info")
    return redirect(url_for("main.criteria_list"))

# ----------------- Subcriteria CRUD -----------------
@main_bp.route("/subcriteria", methods=["GET", "POST"])
@login_required
def subcriteria_list():
    form = SubcriteriaForm()
    form.criteria_id.choices = [
        (c.id, c.name) for c in Criteria.query.order_by(Criteria.display_order.asc()).all()
    ]

    if form.validate_on_submit():
        s = Subcriteria(
            criteria_id=form.criteria_id.data,
            name=form.name.data.strip(),
            min_val=form.min_val.data,
        )
        db.session.add(s)
        db.session.commit()
        flash("Subkriteria ditambahkan.", "success")
        return redirect(url_for("main.subcriteria_list"))

    items = Subcriteria.query.all()
    criteria_lookup = {
        c.id: c.name for c in Criteria.query.order_by(Criteria.display_order.asc()).all()
    }
    return render_template(
        "subcriteria.html",
        form=form,
        items=items,
        criteria_lookup=criteria_lookup,
    )


@main_bp.route("/subcriteria/<int:sid>/edit", methods=["GET", "POST"])
@login_required
def subcriteria_edit(sid):
    s = Subcriteria.query.get_or_404(sid)

    form = SubcriteriaForm(obj=s)
    form.criteria_id.choices = [
        (c.id, c.name) for c in Criteria.query.order_by(Criteria.display_order.asc()).all()
    ]

    if request.method == "GET":
        form.criteria_id.data = s.criteria_id

    if form.validate_on_submit():
        s.criteria_id = form.criteria_id.data
        s.name = form.name.data.strip()
        s.min_val = form.min_val.data
        db.session.commit()
        flash("Subkriteria diupdate.", "success")
        return redirect(url_for("main.subcriteria_list"))

    return render_template("subcriteria_edit.html", form=form, s=s)


@main_bp.route("/subcriteria/<int:sid>/delete", methods=["POST"])
@login_required
def subcriteria_delete(sid):
    s = Subcriteria.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()
    flash("Subkriteria dihapus.", "info")
    return redirect(url_for("main.subcriteria_list"))


# --------------- ARAS Matrix & Ranking ----------------
@main_bp.route("/aras/matrix")
@login_required
def aras_matrix():
    up = _latest_upload()
    if not up or not up.processed_path or not os.path.exists(up.processed_path):
        flash("Belum ada data terproses. Silakan upload data.", "warning")
        return redirect(url_for("main.upload"))

    df = pd.read_csv(up.processed_path)

    # Ambil kriteria & tipe, dan bobot ROC
    criteria_items = Criteria.query.order_by(Criteria.display_order.asc()).all()
    if not criteria_items:
        flash("Belum ada kriteria. Tambahkan dulu di menu Kriteria.", "warning")
        return redirect(url_for("main.criteria_list"))

    criteria = [c.name for c in criteria_items]
    types = [c.ctype for c in criteria_items]

    roc = roc_weights(len(criteria))
    weights = pd.Series(roc, index=criteria)

    # Matriks A & normalisasi (termasuk A0)
    df_all, df_norm = aras_compute(df, criteria, types)

    # Matriks ternormalisasi * bobot
    df_weighted = df_norm[['Alternatif'] + criteria].copy()
    for col in criteria:
        df_weighted[col] = df_weighted[col] * weights[col]

    # Hitung S_i & U_i (pakai util supaya konsisten dengan ranking)
    df_norm_scored, _, _ = aras_score_and_rank(df_all, df_norm, weights, criteria, types)

    # Tabel bobot
    weights_df = pd.DataFrame({
        "Kriteria": criteria,
        "Tipe": types,
        "Bobot ROC": [weights[c] for c in criteria],
    })

    return render_template(
        "aras_matrix.html",
        columns=df_all.columns.tolist(),
        # preview existing (tetap tampil)
        sample_all=df_all.head(20).to_html(classes="table table-sm table-striped", index=False),
        sample_norm=df_norm.head(20).to_html(classes="table table-sm table-striped", index=False),
        # new sections
        weights_html=weights_df.to_html(classes="table table-sm table-striped", index=False, float_format="%.6f"),
        weighted_html=df_weighted.head(20).to_html(classes="table table-sm table-striped", index=False, float_format="%.6f"),
        si_ui_html=df_norm_scored[['Alternatif', 'S_i', 'U_i']].head(20).to_html(classes="table table-sm table-striped", index=False, float_format="%.6f"),
    )

@main_bp.route("/aras/ranking", methods=["GET", "POST"])
@login_required
def aras_ranking():
    up = _latest_upload()
    if not up or not up.processed_path or not os.path.exists(up.processed_path):
        flash("Belum ada data terproses. Silakan upload data.", "warning")
        return redirect(url_for("main.upload"))
    df = pd.read_csv(up.processed_path)

    # ===== Helper: cari kolom fleksibel (abaikan spasi/underscore, case-insensitive)
    def _norm(s: str) -> str:
        return str(s).strip().lower().replace(" ", "").replace("_", "")

    def _find_col(frame, wanted):
        norm_map = {_norm(c): c for c in frame.columns}
        for w in wanted:
            if w in norm_map:
                return norm_map[w]
        return None

    col_jur = _find_col(df, ["jurusan"])
    col_per = _find_col(df, ["periode"])
    col_jp  = _find_col(df, [
        "jurusanperiode", "jurusanperiodegabung", "jurusanperiodecombined",
        "jurusan&periode", "jurusandanperiode", "jurusan-periode"
    ])

    # Pakai mode gabungan kalau kolom gabungan ada dan kolom terpisah tidak ada
    use_combined = col_jp is not None and (col_jur is None and col_per is None)

    # ==== Dropdown values & selected ====
    if use_combined:
        jurusanperiode_values = sorted(df[col_jp].dropna().astype(str).unique().tolist())
        selected_jurusanperiode = request.form.get("jurusanperiode") if request.method == "POST" else ""
    else:
        jurusan_values = sorted(df[col_jur].dropna().astype(str).unique().tolist()) if col_jur else []
        periode_values = sorted(df[col_per].dropna().astype(str).unique().tolist()) if col_per else []
        selected_jurusan = request.form.get("jurusan") if request.method == "POST" else ""
        selected_periode = request.form.get("periode") if request.method == "POST" else ""

    # ==== Filtering ====
    dff = df.copy()
    if use_combined:
        if selected_jurusanperiode:
            dff = dff[dff[col_jp].astype(str) == selected_jurusanperiode]
    else:
        if col_jur and selected_jurusan:
            dff = dff[dff[col_jur].astype(str) == selected_jurusan]
        if col_per and selected_periode:
            dff = dff[dff[col_per].astype(str) == selected_periode]

    # ==== Hitung ROC–ARAS ====
    criteria_items = Criteria.query.order_by(Criteria.display_order.asc()).all()
    criteria = [c.name for c in criteria_items]
    types = [c.ctype for c in criteria_items]
    roc = roc_weights(len(criteria)) if criteria else []
    weights = pd.Series(roc, index=criteria) if criteria else pd.Series(dtype=float)

    df_all, df_norm = aras_compute(dff, criteria, types)
    df_norm, hasil_siswa, _ = aras_score_and_rank(df_all, df_norm, weights, criteria, types)

    # ====== Susun DataFrame rapi untuk CSV ======
    result_df = hasil_siswa.copy()

    # Tambahkan kolom identitas filter yang ada
    extra_cols = []
    for c in [col_jur, col_per, col_jp]:
        if c and c in dff.columns:
            extra_cols.append(c)
    if extra_cols:
        extras = dff[["rid"] + extra_cols].drop_duplicates(subset=["rid"])
        result_df = result_df.merge(extras, on="rid", how="left")

    # Urut kolom rapi
    preferred = ["Ranking", "Alternatif", "S_i", "U_i"]
    if use_combined and col_jp:
        preferred.insert(2, col_jp)
    else:
        if col_jur: preferred.insert(2, col_jur)
        if col_per: preferred.insert(3, col_per)
    criteria_cols = [c for c in criteria if c in result_df.columns]
    other_cols = [c for c in result_df.columns if c not in (preferred + criteria_cols + ["rid"])]
    result_df = result_df[[c for c in preferred if c in result_df.columns] + criteria_cols + other_cols]

    for c in ["S_i", "U_i"]:
        if c in result_df.columns:
            result_df[c] = result_df[c].astype(float).round(6)

    # ====== Nama file dinamis (ikut filter) ======
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    from werkzeug.utils import secure_filename

    parts = ["ranking"]
    if use_combined:
        if selected_jurusanperiode: parts.append(selected_jurusanperiode)
    else:
        if col_jur and selected_jurusan: parts.append(selected_jurusan)
        if col_per and selected_periode: parts.append(selected_periode)

    base = secure_filename("_".join(parts)) or "ranking"
    csv_filename = f"{base}_{ts}.csv"
    excel_filename = f"{base}_{ts}.xlsx"
    csv_path = os.path.join(current_app.config["DOWNLOAD_FOLDER"], csv_filename)
    excel_path = os.path.join(current_app.config["DOWNLOAD_FOLDER"], excel_filename)

    # ====== Tulis CSV rapi ======
    result_df.to_csv(csv_path, index=False, encoding="utf-8-sig", lineterminator="\n")
    download_csv_url = url_for("main.download_file", filename=csv_filename)

    # ====== Tulis Excel rapi ======
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        result_df.to_excel(writer, sheet_name="Ranking", index=False)

        # Sheet Info
        info_rows = [["Dibuat pada", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]]
        if use_combined:
            info_rows.append(["Mode Filter", "JurusanPeriode"])
            info_rows.append(["Filter JurusanPeriode", selected_jurusanperiode or "(Semua)"])
        else:
            info_rows.append(["Mode Filter", "Jurusan + Periode"])
            info_rows.append(["Filter Jurusan", selected_jurusan or "(Semua)"])
            info_rows.append(["Filter Periode", selected_periode or "(Semua)"])
        pd.DataFrame(info_rows, columns=["Kunci", "Nilai"]).to_excel(writer, sheet_name="Info", index=False)

        if len(criteria) > 0:
            pd.DataFrame({
                "Kriteria": criteria,
                "Tipe": types,
                "Bobot ROC": [weights[c] for c in criteria],
            }).to_excel(writer, sheet_name="Bobot ROC", index=False)

        ws = writer.sheets["Ranking"]
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="DDDDDD")
        header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        headers = [c.value for c in ws[1]]
        def col_letter(col_name):
            if col_name in headers:
                return get_column_letter(headers.index(col_name) + 1)
            return None
        col_Si = col_letter("S_i"); col_Ui = col_letter("U_i"); col_Rank = col_letter("Ranking")
        max_row = ws.max_row
        if col_Si:  [setattr(cell, "number_format", "0.000000") for cell in ws[f"{col_Si}2:{col_Si}{max_row}"][0]]
        if col_Ui:  [setattr(cell, "number_format", "0.000000") for cell in ws[f"{col_Ui}2:{col_Ui}{max_row}"][0]]
        if col_Rank:[setattr(cell, "number_format", "0")        for cell in ws[f"{col_Rank}2:{col_Rank}{max_row}"][0]]

        for col_idx, _ in enumerate(ws.columns, start=1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter]:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                except:
                    pass
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 50)

    # ==== Render sesuai mode ====
    if use_combined:
        return render_template(
            "ranking.html",
            use_combined=True,
            jurusanperiode_values=jurusanperiode_values,
            selected_jurusanperiode=selected_jurusanperiode,
            table_html=hasil_siswa.to_html(classes="table table-sm table-hover", index=False),
            download_url=download_csv_url,
            download_excel_url=url_for("main.download_file", filename=excel_filename),
        )
    else:
        return render_template(
            "ranking.html",
            use_combined=False,
            jurusan_values=jurusan_values if col_jur else [],
            periode_values=periode_values if col_per else [],
            selected_jurusan=selected_jurusan if col_jur else "",
            selected_periode=selected_periode if col_per else "",
            table_html=hasil_siswa.to_html(classes="table table-sm table-hover", index=False),
            download_url=download_csv_url,
            download_excel_url=url_for("main.download_file", filename=excel_filename),
        )


@main_bp.route("/downloads/<path:filename>")
@login_required
def download_file(filename):
    path = os.path.join(current_app.config["DOWNLOAD_FOLDER"], filename)
    if not os.path.exists(path):
        flash("File tidak ditemukan.", "warning")
        return redirect(url_for("main.dashboard"))
    return send_file(path, as_attachment=True)
